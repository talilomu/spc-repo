import glob
import tabula
import pandas as pd
import openpyxl
from datetime import datetime
from pdfminer.high_level import extract_text

class ExtractPDF:
    def __init__(self):
        self.path_xl = input("Veuillez renseigner le chemin du fichier Excel...\n")
        self.excel = pd.read_excel(self.path_xl, sheet_name="Feuil1", header=4)
        self.wb = openpyxl.load_workbook(self.path_xl)
        self.feuil = self.wb['Feuil1']
        self.dict_i = {
            'Aspect': ['X', 2],
            'Odeur': ['X', 2],
            'Conductivité à 25 °C': ['AA', 2],
            #'Température de mesure de la conductivité (correction par compensation)': ['', 2],
            'Couleur par Méth. comparative visuelle': ['Y', 2],
            'Couleur': ['Y', 2],
            'pH': ['Z', 2],
            #'Température de mesurage du pH': ['', 2],
            'Turbidité par néphélométrie': ['AB', 2],
            'Turbidité': ['AB', 2],
            'Dénombrement  des micro organismes revivifiables à 37°C': ['AD', 2],
            'Micro-organismes revivifiables 36°C / 48 h': ['AD', 2],
            'Recherche et dénombrement  des bactéries coliformes': ['AE', 2],
            'Coliformes': ['AE', 2],
            'Recherche et dénombrement des Escherichia Coli': ['AF', 2],
            'Escherichia coli': ['AF', 2],
            #'Recherche et dénombrement des spores de micro-organismes'
            #'anaérobies sulfito-réducteurs (Clostridia)': ['', 2],
            'Dénombrement des enterocoques': ['AG', 2],
            'Entérocoques': ['AG', 2],
            'Aluminium (sous-traitance)': ['AS', 2],
            'Aluminium (Al)': ['AS', 2],
            'Fer (sous-traitance)': ['AZ', 2],
            'Fer (Fe)': ['AZ', 2],
            'Plomb (sous-traitance)': ['CK', 2],
            'Plomb (Pb)': ['CK', 2],
            'Zinc (sous-traitance)': ['BB', 2],
            'Zinc (Zn)': ['BB', 2],
            'Nitrates par chromatographie ionique': ['AU', 2],
            'Nitrates (NO3-)': ['AU', 2],
            'Nitrites par chromatographie ionique': ['AV', 2],
            'Nitrites (NO2-)': ['AV', 2],
            #'Matières en suspension': ['', 2],
            'Ammoniums par absorption moléculaire': ['AW', 2],
            'Ammonium (NH4+)': ['AW', 2],
            'Chlore libre mesuré par le client': ['AC', 2],
            'Chlore Libre (Cl2)': ['AC', 2],
            'Température de prélèvement / collecte': ['O', 2],
            'Température de réception': ['R', 2],
            #'Condition de stockage': ['', 1],
            'Hydrogène sulfuré': ['AY', 2],
            #'Température de mesure pour le pH': ['', 2]
            'Oxygène dissous': ['BH', 2],
            'Résidus secs à 180°C': ['AT', 2],
            'Anhydride Carbonique': ['BL', 2],
            'Carbonates (CaCO3)': ['BJ', 2],
            'Hydrogénocarbonates': ['BK', 2],
            'Oxydabilité au KMnO4': ['AX', 2],
            'Phosphore Total (P2O5)': ['BD', 2],
            'Fluorures (F-)': ['BE', 2],
            'Chlorures (Cl-)': ['AN', 2],
            'Sulfates (SO42-)': ['AP', 2],
            'Silice (SiO2)': ['BF', 2],
            'Calcium (Ca)': ['BG', 2],
            'Magnésium (Mg)': ['AQ', 2],
            'Sodium (Na)': ['AO', 2],
            'Potassium (K)': ['AR', 2],
            'Cuivre (Cu)': ['BA', 2],
            'Manganèse (Mn)': ['BC', 2],
            'Cadmium (Cd)': ['CJ', 2],
            'Benzo (a) pyrène': ['CO', 2],
            'Benzo (b) fluoranthène': ['CM', 2],
            'Benzo (g,h,i) pérylène': ['CP', 2],
            'Benzo (k) fluoranthène': ['CN', 2],
            'Fluoranthène': ['CL', 2],
            'Indéno (1,2,3-cd) pyrène': ['CQ', 2],
            'Somme des 6 HPA': ['CR', 2],
            'Naphtalène': ['CS', 2],
            'Acénaphtène': ['CU', 2],
            'Acénaphtylène': ['CT', 2],
            'Anthracène': ['CX', 2],
            'Benzo (a) anthracène': ['CZ', 2],
            'Chrysène': ['DA', 2],
            'Dibenzo (a,h) anthracène': ['DB', 2],
            'Fluorène': ['CV', 2],
            'Pyrène': ['CY', 2],
            'Phénanthrène': ['CW', 2]
        }

    def extrac_rapport(self, f):

        text = extract_text(f)
        lines = text.splitlines()
        df_ls = tabula.read_pdf(f, pages="all")
        today = datetime.today().strftime("%d/%m/%Y")

        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        if(lines[0].strip() == "Institut Louis MALARDE"):
            print("++++++++ RAPPORT D'ANALYSES ILM ++++++++")
            print(lines[8])
            ref_rap = lines[8].split("n°")[1]
            print(lines[10])
            commune = lines[10].split()[2].capitalize()
            for i in range(len(lines)):
                if(lines[i].strip() == "Déposé le"):
                    print(lines[i], lines[i + 4])
                    dt_depot = lines[i + 4].split()[1]
                    hr_depot = lines[i + 4].split()[2]
                elif(lines[i].strip() == "Prélevé le"):
                    print(lines[i], lines[i + 8], lines[i + 10])
                    dt_prel = lines[i + 8].split()[1]
                    hr_prel = lines[i + 8].split()[2]
                    par = lines[i + 10]
                elif(lines[i].strip() == "Température de réception (°C)"):
                    print(lines[i], lines[i + 2])
                    temp_recp = lines[i + 2].strip(":").strip("°C")
                elif (lines[i].strip() == "Méthode de prélèvement"):
                    print(lines[i], lines[i + 2])
                    methode = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Nature échantillon"):
                    print(lines[i], lines[i + 2])
                    nat_ech = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Type d'analyse"):
                    print(lines[i], lines[i + 2])
                    type_anl = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Commune du point"):
                    print(lines[i], lines[i + 2])
                    com_pt = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Nom du point"):
                    print(lines[i], lines[i + 2])
                    nm_pt = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Localisation du point"):
                    print(lines[i], lines[i + 2])
                    loc_pt = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Date début d'analyse"):
                    print(lines[i], lines[i + 2])
                    dt_anl = lines[i + 2].strip(":")
                    hr_anl = lines[i + 2].split("à")[1]
                elif (lines[i].strip().startswith("Observation(s) terrain :")):
                    terrain = []
                    while lines[i].strip().startswith("Observation(s) échantillon :") is False:
                        terrain.append(lines[i].strip())
                        i += 1
                    terrain = [elem for elem in terrain if elem != '']
                    terrain = " ".join(terrain)
                    print(terrain)
                elif (lines[i].strip().startswith("Observation(s) échantillon :")):
                    obs_echan = []
                    while lines[i].strip().startswith("Conclusion Chimie") is False:
                        obs_echan.append(lines[i])
                        i += 1
                    obs_echan = [elem for elem in obs_echan if elem != '']
                    obs_echan = " ".join(obs_echan)
                    print(obs_echan)
                elif (lines[i].strip().startswith("Conclusion Chimie")):
                    chimie = []
                    while lines[i].strip().startswith("Conclusion Bactériologie") is False:
                        chimie.append(lines[i])
                        i += 1
                    chimie = [elem for elem in chimie if elem != '']
                    chimie = " ".join(chimie)
                    print(chimie)
                elif (lines[i].strip().startswith("Conclusion Bactériologie")):
                    bacterio = []
                    while lines[i].strip() != "":
                        bacterio.append(lines[i])
                        i += 1
                    bacterio = [elem for elem in bacterio if elem != '']
                    bacterio = " ".join(bacterio)
                    print(bacterio)
            print("=" * 70)
            new_r = self.feuil.max_row + 1
            self.feuil['A' + str(new_r)] = 'AUTO'
            self.feuil['B' + str(new_r)] = today
            self.feuil['C' + str(new_r)] = commune
            self.feuil['D' + str(new_r)] = nm_pt
            self.feuil['E' + str(new_r)] = loc_pt
            self.feuil['F' + str(new_r)] = dt_prel
            self.feuil['G' + str(new_r)] = ref_rap
            self.feuil['H' + str(new_r)] = 'Com.'
            self.feuil['I' + str(new_r)] = 'Autocontrôle'
            self.feuil['J' + str(new_r)] = type_anl
            self.feuil['K' + str(new_r)] = nat_ech
            self.feuil['L' + str(new_r)] = '???'
            self.feuil['M' + str(new_r)] = 'Commune'
            self.feuil['N' + str(new_r)] = hr_prel
            self.feuil['O' + str(new_r)] = temp_recp
            self.feuil['P' + str(new_r)] = dt_depot
            self.feuil['Q' + str(new_r)] = hr_depot
            self.feuil['R' + str(new_r)] = '???'
            self.feuil['S' + str(new_r)] = dt_anl
            self.feuil['T' + str(new_r)] = hr_anl
            self.feuil['U' + str(new_r)] = 'ILM'
            for dt in df_ls:
                num_unnamed = sum(["Unnamed" in col for col in dt.columns])
                if (len(dt.columns) >= 5):
                    if(num_unnamed < 2):
                        header_values = dt.columns.values.tolist()
                        for index, value in enumerate(self.dict_i):
                            if(header_values[0] == value):
                                self.feuil[self.dict_i[value][0] + str(new_r)] = header_values[self.dict_i[value][1]]
                    for i, row in dt.iterrows():
                        for index, value in enumerate(self.dict_i):
                            if(row.iloc[0] == value):
                                self.feuil[self.dict_i[value][0] + str(new_r)] = row.iloc[self.dict_i[value][1]]

        else:
            print("++++++++ RAPPORT D'ESSAI CAIRAP ++++++++")
            print(lines[3], lines[4])
            ref_rap = lines[4]
            for i in range(len(lines)):
                if (lines[i].strip().startswith("COMMUNE DE")):
                    commune = lines[i].strip().split()[2].capitalize()
                    print("Commune de " + commune)
                elif (lines[i].strip().startswith("987")):
                    lieu_prel = lines[i].split()[-1].capitalize()
                elif (lines[i].strip() == "Lieu de prélèvement (#)"):
                    print(lines[i], lines[i + 2])
                    nm_pt = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Identification échantillon (#)"):
                    print(lines[i], lines[i + 2])
                    type_anl = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Nature échantillon (#)"):
                    print(lines[i], lines[i + 2])
                    nat_ech = lines[i + 2].strip(":")
                elif (lines[i].strip() == "Echantillon prélevé le"):
                    print(lines[i], lines[i + 2])
                    dt_prel = lines[i + 2].split()[1]
                    hr_prel = lines[i + 2].split()[3]
                elif (lines[i].strip() == "Echantillon réceptionné le"):
                    print(lines[i], lines[i + 2])
                    dt_depot = lines[i + 2].split()[1]
                    hr_depot = lines[i + 2].split()[3]
                elif (lines[i].strip() == "Echantillon analysé le"):
                    print(lines[i], lines[i + 2])
                    dt_anl = lines[i + 2].split()[1]
                    hr_anl = lines[i + 2].split()[3]
                elif (lines[i].strip() == "Observation échantillon"):
                    obs_echan = []
                    while lines[i].strip() != "Température de prélèvement / collecte":
                        obs_echan.append(lines[i])
                        i += 1
                    obs_echan = [elem for elem in obs_echan if elem != '']
                    obs_echan = " ".join(obs_echan)
                    print(obs_echan)
                elif (lines[i].strip() == "Déclaration de conformité :"):
                    print(lines[i], lines[i + 1])
            print("=" * 70)
            new_r = self.feuil.max_row + 1
            self.feuil['A' + str(new_r)] = 'AUTO'
            self.feuil['B' + str(new_r)] = today
            self.feuil['C' + str(new_r)] = commune
            self.feuil['D' + str(new_r)] = nm_pt
            self.feuil['E' + str(new_r)] = lieu_prel
            self.feuil['F' + str(new_r)] = dt_prel
            self.feuil['G' + str(new_r)] = ref_rap
            self.feuil['H' + str(new_r)] = 'Com.'
            self.feuil['I' + str(new_r)] = 'Autocontrôle'
            self.feuil['J' + str(new_r)] = type_anl
            self.feuil['K' + str(new_r)] = nat_ech
            self.feuil['L' + str(new_r)] = '???'
            self.feuil['M' + str(new_r)] = 'Commune'
            self.feuil['N' + str(new_r)] = hr_prel
            #self.feuil['O' + str(new_r)] = temp_recp
            self.feuil['P' + str(new_r)] = dt_depot
            self.feuil['Q' + str(new_r)] = hr_depot
            self.feuil['R' + str(new_r)] = '???'
            self.feuil['S' + str(new_r)] = dt_anl
            self.feuil['T' + str(new_r)] = hr_anl
            self.feuil['U' + str(new_r)] = 'CAIRAP'
            for dt in df_ls:
                num_unnamed = sum(["Unnamed" in col for col in dt.columns])
                if(len(dt.columns) == 3):
                    if (num_unnamed < 2):
                        header_values = dt.columns.values.tolist()
                        for index, value in enumerate(self.dict_i):
                            if(header_values[0] == value):
                                self.feuil[self.dict_i[value][0] + str(new_r)] = header_values[self.dict_i[value][1]]
                    for i, row in dt.iterrows():
                        print(row.iloc[0], row.iloc[2])
                        for index, value in enumerate(self.dict_i):
                            if (row.iloc[0] == value):
                                self.feuil[self.dict_i[value][0] + str(new_r)] = row.iloc[self.dict_i[value][1]]
                elif(len(dt.columns) == 6):
                    if (num_unnamed < 2):
                        header_values = dt.columns.values.tolist()
                        for index, value in enumerate(self.dict_i):
                            if (header_values[0] == value):
                                self.feuil[self.dict_i[value][0] + str(new_r)] = header_values[self.dict_i[value][1]]
                    for i, row in dt.iterrows():
                        print(row.iloc[0], row.iloc[2])
                        for index, value in enumerate(self.dict_i):
                            if(row.iloc[0] == value):
                                self.feuil[self.dict_i[value][0] + str(new_r)] = row.iloc[self.dict_i[value][1]]


        self.wb.save(self.path_xl)
        self.wb.close()


if __name__ == '__main__':

    files = glob.glob(input("Veuillez indiquer le chemin absolu du dossier contenant les analyses...\n") + "\*.pdf")
    start = ExtractPDF()
    for f in files:
       start.extrac_rapport(f)
